import os
from pathlib import Path
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
import xml.etree.ElementTree as ET


SHEET_FAKTUR = "Faktur"
SHEET_DETAIL = "DetailFaktur"

FAKTUR_HEADER_ROW = 3
DETAIL_HEADER_ROW = 1
JOIN_KEY_COL_NAME = "Baris"

FAKTUR_COL_TO_XML = {
    "Tanggal Faktur": "TaxInvoiceDate",
    "Jenis Faktur": "TaxInvoiceOpt",
    "Kode Transaksi": "TrxCode",
    "Keterangan Tambahan": "AddInfo",
    "Dokumen Pendukung": "CustomDoc",
    "Period Dok Pendukung": "CustomDocMonthYear",
    "Referensi": "RefDesc",
    "Cap Fasilitas": "FacilityStamp",
    "ID TKU Penjual": "SellerIDTKU",
    "NPWP/NIK Pembeli": "BuyerTin",
    "Jenis ID Pembeli": "BuyerDocument",
    "Negara Pembeli": "BuyerCountry",
    "Nomor Dokumen Pembeli": "BuyerDocumentNumber",
    "Nama Pembeli": "BuyerName",
    "Alamat Pembeli": "BuyerAdress",
    "Email Pembeli": "BuyerEmail",
    "ID TKU Pembeli": "BuyerIDTKU",
}

DETAIL_COL_TO_XML = {
    "Barang/Jasa": "Opt",
    "Kode Barang Jasa": "Code",
    "Nama Barang/Jasa": "Name",
    "Nama Satuan Ukur": "Unit",
    "Harga Satuan": "Price",
    "Jumlah Barang Jasa": "Qty",
    "Total Diskon": "TotalDiscount",
    "DPP": "TaxBase",
    "DPP Nilai Lain": "OtherTaxBase",
    "Tarif PPN": "VATRate",
    "PPN": "VAT",
    "Tarif PPnBM": "STLGRate",
    "PPnBM": "STLG",
}


def read_sheet(ws, header_row, end_marker="END"):
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    headers = [str(c).strip() if c is not None else "" for c in header_cells]

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        first_col = row[0]
        if first_col is None or (isinstance(first_col, str) and first_col.strip().upper() == end_marker):
            break
        d = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(d)
    return rows


def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    return s


def format_numeric(val):
    if val is None or val == "":
        return ""
    try:
        num = Decimal(str(val).replace(",", "."))
        rounded = num.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{rounded:.2f}"
    except:
        return str(val)


def get_tin(faktur):
    for r in faktur:
        v = r.get("ID TKU Penjual")
        if v:
            s = str(v).strip()
            return s[:16] if len(s) >= 16 else s
    return ""


def build_xml(faktur, detail):
    detail_by = {}
    for d in detail:
        key = str(d.get(JOIN_KEY_COL_NAME, "")).strip()
        detail_by.setdefault(key, []).append(d)

    root = ET.Element("TaxInvoiceBulk")
    ET.SubElement(root, "TIN").text = get_tin(faktur)
    list_el = ET.SubElement(root, "ListOfTaxInvoice")

    for f in faktur:
        baris = str(f.get(JOIN_KEY_COL_NAME, "")).strip()
        inv = ET.SubElement(list_el, "TaxInvoice")

        for col, tag in FAKTUR_COL_TO_XML.items():
            v = f.get(col)
            text = format_date(v) if tag == "TaxInvoiceDate" else ("" if v is None else str(v))
            ET.SubElement(inv, tag).text = text

        list_gs = ET.SubElement(inv, "ListOfGoodService")
        for d in detail_by.get(baris, []):
            gs = ET.SubElement(list_gs, "GoodService")
            for col, tag in DETAIL_COL_TO_XML.items():
                v = d.get(col)
                if tag in ["Price", "Qty", "TotalDiscount", "TaxBase", "OtherTaxBase", "VATRate", "VAT", "STLGRate", "STLG"]:
                    ET.SubElement(gs, tag).text = format_numeric(v)
                else:
                    ET.SubElement(gs, tag).text = "" if v is None else str(v)

    return ET.ElementTree(root)


def run():
    print("==============================================")
    print("   CONVERTER FAKTUR CORE-TAX (Drag & Drop)")
    print("==============================================\n")
    print("Silakan drag & drop file Excel Faktur di sini lalu tekan ENTER:")
    excel_path = input(">> ").strip().strip('"')

    if not excel_path.lower().endswith(".xlsx"):
        print("File bukan XLSX. Harap input template Excel Faktur DJP.")
        os.system("pause")
        return

    path = Path(excel_path)
    xml_path = path.with_suffix(".xml")

    try:
        wb = load_workbook(path, data_only=True)
        faktur = read_sheet(wb[SHEET_FAKTUR], FAKTUR_HEADER_ROW)
        detail = read_sheet(wb[SHEET_DETAIL], DETAIL_HEADER_ROW)
        xml = build_xml(faktur, detail)
        xml.write(xml_path, encoding="utf-8", xml_declaration=True)
        print("\n==============================================")
        print(" XML berhasil dibuat:")
        print(" -->", xml_path)
        print("==============================================")
    except Exception as e:
        print("Terjadi error:", str(e))

    os.system("pause")


if __name__ == "__main__":
    run()
